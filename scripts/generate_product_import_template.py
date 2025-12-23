#!/usr/bin/env python3
"""
Script para generar template de Excel para carga masiva de productos
Genera un archivo Excel con columnas completas y 3 ejemplos de productos
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json

# Configuración de estilos
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
REQUIRED_FILL = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
OPTIONAL_FILL = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def create_excel_template():
    """Crea el archivo Excel con el template de carga masiva"""
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Carga Masiva Productos"
    
    # Definir columnas
    columns = [
        # Información básica
        ("name", "Nombre del Producto", True, "Ej: Filtro de Aire Original"),
        ("sku", "SKU (Código)", False, "Ej: FIL-AIR-001"),
        ("description", "Descripción", False, "Descripción detallada del producto"),
        ("image_url", "URL de Imagen", False, "https://ejemplo.com/imagen.jpg"),
        ("price", "Precio Base", True, "150.00"),
        ("product_type", "Tipo de Producto", True, "refaccion|accesorio|servicio_instalacion|servicio_mantenimiento|fluido"),
        ("category_id", "ID de Categoría (UUID)", False, "00000001-0000-0000-0000-000000000001"),
        ("is_available", "Disponible", False, "true|false (default: true)"),
        ("is_featured", "Destacado", False, "true|false (default: false)"),
        ("display_order", "Orden de Visualización", False, "0 (default: 0)"),
        
        # Variantes - Grupo 1
        ("variant_group_1_name", "Variante Grupo 1 - Nombre", False, "Ej: Tamaño, Capacidad, Color"),
        ("variant_group_1_required", "Variante Grupo 1 - Requerido", False, "true|false"),
        ("variant_group_1_selection", "Variante Grupo 1 - Tipo Selección", False, "single|multiple"),
        ("variant_group_1_variants", "Variante Grupo 1 - Variantes (JSON)", False, 'Ver formato en instrucciones'),
        
        # Variantes - Grupo 2
        ("variant_group_2_name", "Variante Grupo 2 - Nombre", False, "Ej: Marca, Modelo"),
        ("variant_group_2_required", "Variante Grupo 2 - Requerido", False, "true|false"),
        ("variant_group_2_selection", "Variante Grupo 2 - Tipo Selección", False, "single|multiple"),
        ("variant_group_2_variants", "Variante Grupo 2 - Variantes (JSON)", False, 'Ver formato en instrucciones'),
        
        # Información técnica (nutritional_info como JSON)
        ("technical_specs", "Especificaciones Técnicas (JSON)", False, '{"marca": "Toyota", "modelo": "Corolla", "año": "2020-2023"}'),
    ]
    
    # Escribir encabezados
    for col_idx, (field, header, required, example) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL if not required else HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
        
        # Ajustar ancho de columna
        ws.column_dimensions[get_column_letter(col_idx)].width = 25
    
    # Escribir ejemplos de ayuda en fila 2
    for col_idx, (field, header, required, example) in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=example)
        cell.font = Font(size=9, italic=True, color="666666")
        cell.fill = REQUIRED_FILL if required else OPTIONAL_FILL
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell.border = BORDER
    
    # Ejemplos de productos
    examples = [
        {
            "name": "Filtro de Aire Original Toyota",
            "sku": "FIL-AIR-TOY-001",
            "description": "Filtro de aire original Toyota para modelos Corolla 2020-2023. Filtración eficiente de partículas.",
            "image_url": "https://ejemplo.com/filtro-aire-toyota.jpg",
            "price": "150.00",
            "product_type": "refaccion",
            "category_id": "",
            "is_available": "true",
            "is_featured": "true",
            "display_order": "1",
            "variant_group_1_name": "Compatibilidad",
            "variant_group_1_required": "true",
            "variant_group_1_selection": "single",
            "variant_group_1_variants": json.dumps([
                {"name": "Corolla 2020-2021", "price_adjustment": 0, "is_available": True},
                {"name": "Corolla 2022-2023", "price_adjustment": 20, "is_available": True},
                {"name": "Camry 2020-2023", "price_adjustment": 30, "is_available": True}
            ], ensure_ascii=False),
            "variant_group_2_name": "",
            "variant_group_2_required": "",
            "variant_group_2_selection": "",
            "variant_group_2_variants": "",
            "technical_specs": json.dumps({
                "marca": "Toyota",
                "modelo_compatible": "Corolla, Camry",
                "años": "2020-2023",
                "tipo_filtro": "Aire",
                "material": "Papel sintético"
            }, ensure_ascii=False)
        },
        {
            "name": "Aceite Motor 5W-30 Sintético",
            "sku": "ACE-5W30-SYN-001",
            "description": "Aceite de motor sintético 5W-30 de alto rendimiento. Protección superior del motor.",
            "image_url": "https://ejemplo.com/aceite-5w30.jpg",
            "price": "450.00",
            "product_type": "fluido",
            "category_id": "",
            "is_available": "true",
            "is_featured": "false",
            "display_order": "2",
            "variant_group_1_name": "Capacidad",
            "variant_group_1_required": "true",
            "variant_group_1_selection": "single",
            "variant_group_1_variants": json.dumps([
                {"name": "1 Litro", "price_adjustment": 0, "is_available": True},
                {"name": "4 Litros", "price_adjustment": 1350, "is_available": True},
                {"name": "5 Litros", "price_adjustment": 1650, "is_available": True}
            ], ensure_ascii=False),
            "variant_group_2_name": "Tipo",
            "variant_group_2_required": "false",
            "variant_group_2_selection": "single",
            "variant_group_2_variants": json.dumps([
                {"name": "Sintético", "price_adjustment": 0, "is_available": True},
                {"name": "Semi-Sintético", "price_adjustment": -50, "is_available": True},
                {"name": "Convencional", "price_adjustment": -100, "is_available": True}
            ], ensure_ascii=False),
            "technical_specs": json.dumps({
                "viscosidad": "5W-30",
                "tipo": "Sintético",
                "capacidad_litros": "1, 4, 5",
                "certificaciones": "API SN Plus, ILSAC GF-6",
                "temperatura_operacion": "-30°C a 40°C"
            }, ensure_ascii=False)
        },
        {
            "name": "Instalación de Sistema de Audio",
            "sku": "SERV-AUDIO-INST-001",
            "description": "Servicio profesional de instalación de sistema de audio completo. Incluye mano de obra y garantía.",
            "image_url": "https://ejemplo.com/instalacion-audio.jpg",
            "price": "1200.00",
            "product_type": "servicio_instalacion",
            "category_id": "",
            "is_available": "true",
            "is_featured": "true",
            "display_order": "3",
            "variant_group_1_name": "Tiempo Estimado",
            "variant_group_1_required": "false",
            "variant_group_1_selection": "single",
            "variant_group_1_variants": json.dumps([
                {"name": "2-3 horas", "price_adjustment": 0, "is_available": True},
                {"name": "4-6 horas", "price_adjustment": 300, "is_available": True},
                {"name": "1 día completo", "price_adjustment": 600, "is_available": True}
            ], ensure_ascii=False),
            "variant_group_2_name": "",
            "variant_group_2_required": "",
            "variant_group_2_selection": "",
            "variant_group_2_variants": "",
            "technical_specs": json.dumps({
                "tiempo_estimado": "2-6 horas",
                "dificultad": "Media-Alta",
                "herramientas_requeridas": "Destornilladores, alicates, multímetro",
                "garantia": "3 meses",
                "incluye": "Instalación, cableado, configuración básica"
            }, ensure_ascii=False)
        }
    ]
    
    # Escribir ejemplos
    for row_idx, example in enumerate(examples, start=3):
        for col_idx, (field, header, required, example_text) in enumerate(columns, start=1):
            value = example.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            if required and not value:
                cell.fill = REQUIRED_FILL
    
    # Crear hoja de instrucciones
    ws_instructions = wb.create_sheet("INSTRUCCIONES", 0)
    ws_instructions.column_dimensions['A'].width = 100
    
    instructions = [
        "INSTRUCCIONES PARA CARGA MASIVA DE PRODUCTOS",
        "",
        "COLUMNAS REQUERIDAS (marcadas en rojo):",
        "  • name: Nombre del producto (máximo 255 caracteres)",
        "  • price: Precio base del producto (formato: 150.00)",
        "  • product_type: Tipo de producto (valores válidos):",
        "    - refaccion: Refacción (pieza de repuesto)",
        "    - accesorio: Accesorio (personalización)",
        "    - servicio_instalacion: Servicio de Instalación",
        "    - servicio_mantenimiento: Servicio de Mantenimiento",
        "    - fluido: Fluidos y Lubricantes",
        "",
        "COLUMNAS OPCIONALES:",
        "  • sku: Código único del producto (máximo 100 caracteres)",
        "  • description: Descripción del producto",
        "  • image_url: URL de la imagen del producto",
        "  • category_id: UUID de la categoría (si se conoce)",
        "  • is_available: true/false (default: true)",
        "  • is_featured: true/false (default: false)",
        "  • display_order: Número entero (default: 0)",
        "",
        "VARIANTES:",
        "  Las variantes se definen usando grupos. Cada producto puede tener hasta 2 grupos de variantes.",
        "",
        "  Para cada grupo de variantes:",
        "    • variant_group_X_name: Nombre del grupo (ej: 'Tamaño', 'Color', 'Capacidad')",
        "    • variant_group_X_required: true si es obligatorio seleccionar una variante",
        "    • variant_group_X_selection: 'single' para selección única, 'multiple' para múltiple",
        "    • variant_group_X_variants: JSON con array de variantes",
        "",
        "  Formato JSON para variant_group_X_variants:",
        "    [",
        '      {"name": "Variante 1", "price_adjustment": 0, "is_available": true},',
        '      {"name": "Variante 2", "price_adjustment": 50, "is_available": true},',
        '      {"name": "Variante 3", "absolute_price": 200, "is_available": true}',
        "    ]",
        "",
        "  Campos de cada variante:",
        "    • name: Nombre de la variante (requerido)",
        "    • price_adjustment: Ajuste de precio relativo al precio base (default: 0)",
        "    • absolute_price: Precio absoluto (opcional, si se especifica ignora price_adjustment)",
        "    • is_available: true/false (default: true)",
        "",
        "ESPECIFICACIONES TÉCNICAS:",
        "  • technical_specs: JSON con especificaciones técnicas del producto",
        "  • Formato: Objeto JSON con cualquier estructura",
        "  • Ejemplo:",
        '    {"marca": "Toyota", "modelo": "Corolla", "año": "2020-2023"}',
        "",
        "NOTAS IMPORTANTES:",
        "  • NO incluir campos de stock ni relaciones con sucursales",
        "  • Solo se considera el precio base del producto",
        "  • El category_id es opcional pero recomendado",
        "  • Si no se especifica category_id, el producto se puede asignar manualmente después",
        "  • Los UUIDs deben estar en formato: 00000001-0000-0000-0000-000000000001",
        "  • Para productos sin variantes, dejar las columnas de variantes vacías",
        "",
        "EJEMPLOS:",
        "  Ver las filas 3, 4 y 5 de la hoja 'Carga Masiva Productos' para ejemplos completos.",
    ]
    
    for row_idx, instruction in enumerate(instructions, start=1):
        cell = ws_instructions.cell(row=row_idx, column=1, value=instruction)
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
        elif instruction.startswith("  •") or instruction.startswith("    •"):
            cell.font = Font(size=10)
        else:
            cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Guardar archivo
    filename = "template_carga_masiva_productos.xlsx"
    wb.save(filename)
    print(f"✅ Template creado exitosamente: {filename}")
    print(f"   - Hoja 'INSTRUCCIONES': Instrucciones detalladas")
    print(f"   - Hoja 'Carga Masiva Productos': Template con 3 ejemplos")
    print(f"   - Total de columnas: {len(columns)}")
    print(f"   - Ejemplos incluidos: {len(examples)}")

if __name__ == "__main__":
    try:
        create_excel_template()
    except ImportError:
        print("❌ Error: Se requiere la librería 'openpyxl'")
        print("   Instala con: pip install openpyxl")
    except Exception as e:
        print(f"❌ Error al crear template: {e}")


