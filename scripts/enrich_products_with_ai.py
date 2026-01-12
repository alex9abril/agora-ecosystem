#!/usr/bin/env python3
"""
Script para enriquecer productos con IA
Lee un Excel con productos básicos (número de parte, nombre, existencia, precio)
y usa IA para completar información faltante y buscar imágenes.

Requisitos:
    pip install pandas openpyxl openai requests pillow

Uso:
    python scripts/enrich_products_with_ai.py --input productos.xlsx --output productos_completos.xlsx
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import os
import sys
import argparse
import time
from typing import Dict, Optional, List
import requests
from io import BytesIO
from PIL import Image

# Intentar importar OpenAI (opcional)
try:
    from openai import OpenAI
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False
    print("⚠️  OpenAI no está disponible. Instala con: pip install openai")

# Configuración de estilos para Excel
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUCCESS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Tipos de producto válidos
VALID_PRODUCT_TYPES = [
    'refaccion',
    'accesorio',
    'servicio_instalacion',
    'servicio_mantenimiento',
    'fluido'
]

# Mapeo de palabras clave para tipos de producto
PRODUCT_TYPE_KEYWORDS = {
    'refaccion': ['filtro', 'pastilla', 'disco', 'bujía', 'sensor', 'correa', 'manguera', 
                  'amortiguador', 'rotula', 'terminal', 'alternador', 'arrancador', 'batería',
                  'radiador', 'termostato', 'bomba', 'embrague', 'junta', 'componente'],
    'accesorio': ['audio', 'bocina', 'pantalla', 'led', 'alarma', 'cámara', 'spoiler', 
                  'alerón', 'calcomanía', 'vinilo', 'tapete', 'funda', 'organizador',
                  'portaequipaje', 'remolque', 'portabicicleta'],
    'servicio_instalacion': ['instalación', 'instalacion', 'montaje', 'colocación', 'colocacion'],
    'servicio_mantenimiento': ['mantenimiento', 'servicio', 'cambio', 'revisión', 'revision',
                               'alineación', 'alineacion', 'balanceo', 'diagnóstico', 'diagnostico'],
    'fluido': ['aceite', 'líquido', 'liquido', 'refrigerante', 'aditivo', 'lubricante', 'fluido']
}

class ProductEnricher:
    """Clase para enriquecer productos con IA"""
    
    def __init__(self, openai_api_key: Optional[str] = None, unsplash_api_key: Optional[str] = None):
        self.openai_client = None
        self.unsplash_api_key = unsplash_api_key or os.getenv('UNSPLASH_ACCESS_KEY')
        
        if OPENAI_AVAILABLE and openai_api_key:
            self.openai_client = OpenAI(api_key=openai_api_key)
        elif OPENAI_AVAILABLE:
            # Intentar obtener de variable de entorno
            api_key = os.getenv('OPENAI_API_KEY')
            if api_key:
                self.openai_client = OpenAI(api_key=api_key)
    
    def detect_product_type(self, name: str, part_number: str = "") -> str:
        """Detecta el tipo de producto basado en el nombre y número de parte"""
        text = f"{name} {part_number}".lower()
        
        # Contar coincidencias por tipo
        scores = {}
        for product_type, keywords in PRODUCT_TYPE_KEYWORDS.items():
            score = sum(1 for keyword in keywords if keyword in text)
            if score > 0:
                scores[product_type] = score
        
        if scores:
            # Retornar el tipo con mayor score
            return max(scores, key=scores.get)
        
        # Por defecto, asumir refacción
        return 'refaccion'
    
    def enrich_with_ai(self, name: str, part_number: str = "", price: float = 0) -> Dict:
        """Enriquece un producto usando IA"""
        if not self.openai_client:
            return self._enrich_basic(name, part_number, price)
        
        try:
            prompt = f"""Eres un experto en autopartes y productos automotrices. 
Analiza el siguiente producto y proporciona información detallada en formato JSON:

Producto: {name}
Número de Parte: {part_number}
Precio: ${price:.2f}

Proporciona la siguiente información en formato JSON válido:
{{
    "description": "Descripción detallada del producto (2-4 oraciones)",
    "product_type": "refaccion|accesorio|servicio_instalacion|servicio_mantenimiento|fluido",
    "suggested_category": "Categoría sugerida basada en el tipo de producto",
    "technical_specs": {{
        "marca_compatible": "Marcas de vehículos compatibles (si aplica)",
        "modelos_compatibles": "Modelos de vehículos compatibles (si aplica)",
        "años_compatibles": "Rango de años compatibles (si aplica)",
        "especificaciones": "Otras especificaciones técnicas relevantes"
    }},
    "search_keywords": ["palabra1", "palabra2", "palabra3"] para búsqueda de imágenes
}}

Responde SOLO con el JSON, sin texto adicional."""

            response = self.openai_client.chat.completions.create(
                model="gpt-4o-mini",  # Usar modelo más económico
                messages=[
                    {"role": "system", "content": "Eres un experto en autopartes. Responde siempre en formato JSON válido."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.7,
                max_tokens=500
            )
            
            content = response.choices[0].message.content.strip()
            
            # Limpiar el contenido si tiene markdown
            if content.startswith("```json"):
                content = content[7:]
            if content.startswith("```"):
                content = content[3:]
            if content.endswith("```"):
                content = content[:-3]
            content = content.strip()
            
            data = json.loads(content)
            
            # Validar y completar datos
            if 'product_type' not in data or data['product_type'] not in VALID_PRODUCT_TYPES:
                data['product_type'] = self.detect_product_type(name, part_number)
            
            return data
            
        except json.JSONDecodeError as e:
            print(f"⚠️  Error parseando JSON de IA: {e}")
            return self._enrich_basic(name, part_number, price)
        except Exception as e:
            print(f"⚠️  Error con IA: {e}")
            return self._enrich_basic(name, part_number, price)
    
    def _enrich_basic(self, name: str, part_number: str = "", price: float = 0) -> Dict:
        """Enriquecimiento básico sin IA"""
        product_type = self.detect_product_type(name, part_number)
        
        return {
            "description": f"{name}. Producto de calidad para vehículos.",
            "product_type": product_type,
            "suggested_category": "General",
            "technical_specs": {
                "marca_compatible": "",
                "modelos_compatibles": "",
                "años_compatibles": "",
                "especificaciones": ""
            },
            "search_keywords": [name.lower(), part_number] if part_number else [name.lower()]
        }
    
    def search_image_url(self, name: str, part_number: str = "", keywords: List[str] = None, unsplash_api_key: Optional[str] = None) -> Optional[str]:
        """Busca una URL de imagen del producto usando Unsplash API"""
        # Construir términos de búsqueda
        search_terms = []
        if part_number:
            search_terms.append(part_number)
        if name:
            # Limpiar nombre para búsqueda
            clean_name = name.lower()
            # Remover palabras comunes
            stop_words = ['de', 'del', 'la', 'el', 'para', 'con', 'y', 'o']
            words = [w for w in clean_name.split() if w not in stop_words and len(w) > 2]
            search_terms.extend(words[:3])  # Máximo 3 palabras del nombre
        
        if keywords:
            search_terms.extend(keywords[:2])  # Limitar a 2 keywords adicionales
        
        # Construir query de búsqueda
        query = " ".join(search_terms[:5])  # Máximo 5 términos
        
        if not query:
            return None
        
        # Intentar buscar en Unsplash si hay API key
        if unsplash_api_key:
            try:
                url = "https://api.unsplash.com/search/photos"
                headers = {
                    "Authorization": f"Client-ID {unsplash_api_key}"
                }
                params = {
                    "query": query,
                    "per_page": 1,
                    "orientation": "landscape"
                }
                
                response = requests.get(url, headers=headers, params=params, timeout=5)
                if response.status_code == 200:
                    data = response.json()
                    if data.get('results') and len(data['results']) > 0:
                        image_url = data['results'][0]['urls']['regular']
                        return image_url
            except Exception as e:
                print(f"    ⚠️  Error buscando imagen en Unsplash: {e}")
        
        # Si no hay API key o falló, retornar None
        # El usuario puede agregar URLs manualmente
        return None
    
    def format_technical_specs(self, specs: Dict) -> str:
        """Formatea las especificaciones técnicas en formato pipe-separated"""
        if not specs:
            return ""
        
        parts = []
        for key, value in specs.items():
            if value:
                parts.append(f"{key}:{value}")
        
        return "|".join(parts)


def read_input_excel(file_path: str) -> pd.DataFrame:
    """Lee el Excel de entrada y normaliza las columnas"""
    try:
        df = pd.read_excel(file_path)
        
        # Normalizar nombres de columnas (case-insensitive, sin espacios)
        df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
        
        # Mapear posibles nombres de columnas
        column_mapping = {
            'numero_de_parte': 'part_number',
            'número_de_parte': 'part_number',
            'part_number': 'part_number',
            'partnumber': 'part_number',
            'sku': 'part_number',
            'codigo': 'part_number',
            'código': 'part_number',
            'nombre': 'name',
            'name': 'name',
            'producto': 'name',
            'descripcion': 'name',
            'existencia': 'stock',
            'stock': 'stock',
            'inventario': 'stock',
            'cantidad': 'stock',
            'precio': 'price',
            'price': 'price',
            'precio_unitario': 'price',
            'costo': 'price'
        }
        
        # Renombrar columnas
        df = df.rename(columns=column_mapping)
        
        # Validar columnas requeridas
        required_columns = ['name', 'price']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            raise ValueError(f"Faltan columnas requeridas: {', '.join(missing_columns)}")
        
        return df
        
    except Exception as e:
        raise Exception(f"Error leyendo Excel: {e}")


def create_enriched_excel(df: pd.DataFrame, enricher: ProductEnricher, output_path: str, search_images: bool = True):
    """Crea un Excel enriquecido con toda la información"""
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos Enriquecidos"
    
    # Definir columnas del Excel de salida
    columns = [
        ("name", "Nombre del Producto", True),
        ("sku", "SKU (Número de Parte)", False),
        ("description", "Descripción", False),
        ("image_url", "URL de Imagen", False),
        ("price", "Precio Base", True),
        ("product_type", "Tipo de Producto", True),
        ("category_slug", "Slug de Categoría", False),
        ("is_available", "Disponible", False),
        ("is_featured", "Destacado", False),
        ("display_order", "Orden de Visualización", False),
        ("technical_specs", "Especificaciones Técnicas", False),
        ("stock", "Existencia Original", False),
    ]
    
    # Escribir encabezados
    for col_idx, (field, header, required) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = 25
    
    # Procesar cada producto
    total_products = len(df)
    print(f"\n📦 Procesando {total_products} productos...")
    
    for idx, row in df.iterrows():
        product_num = idx + 1
        print(f"  [{product_num}/{total_products}] Procesando: {row.get('name', 'N/A')}")
        
        # Obtener datos básicos
        name = str(row.get('name', '')).strip()
        part_number = str(row.get('part_number', row.get('sku', ''))).strip()
        price = float(row.get('price', 0))
        stock = row.get('stock', row.get('existencia', ''))
        
        # Enriquecer con IA
        enriched_data = enricher.enrich_with_ai(name, part_number, price)
        
        # Buscar imagen (opcional)
        image_url = None
        if search_images:
            image_url = enricher.search_image_url(
                name, 
                part_number, 
                enriched_data.get('search_keywords', []),
                enricher.unsplash_api_key
            )
            if image_url:
                print(f"    ✅ Imagen encontrada")
        
        # Formatear especificaciones técnicas
        tech_specs = enricher.format_technical_specs(enriched_data.get('technical_specs', {}))
        
        # Determinar disponibilidad basada en stock
        is_available = True
        if pd.notna(stock):
            try:
                stock_num = float(stock)
                is_available = stock_num > 0
            except:
                pass
        
        # Escribir fila
        row_num = idx + 2
        values = [
            name,
            part_number if part_number else "",
            enriched_data.get('description', ''),
            image_url if image_url else "",
            price,
            enriched_data.get('product_type', 'refaccion'),
            "",  # category_slug - el usuario puede completarlo
            "true" if is_available else "false",
            "false",
            0,
            tech_specs,
            stock if pd.notna(stock) else ""
        ]
        
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            # Colorear según estado
            if col_idx == 8:  # is_available
                if value == "true":
                    cell.fill = SUCCESS_FILL
                else:
                    cell.fill = WARNING_FILL
    
    # Crear hoja de resumen
    ws_summary = wb.create_sheet("Resumen", 0)
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 50
    
    summary_data = [
        ["RESUMEN DE ENRIQUECIMIENTO", ""],
        ["", ""],
        ["Total de productos procesados:", total_products],
        ["Fecha de procesamiento:", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["", ""],
        ["NOTAS:", ""],
        ["", "• Revisa y completa la columna 'Slug de Categoría' usando catalogo_categorias.csv"],
        ["", "• Agrega URLs de imágenes en la columna 'URL de Imagen' si están disponibles"],
        ["", "• Verifica que los tipos de producto sean correctos"],
        ["", "• Ajusta las descripciones si es necesario"],
        ["", ""],
        ["TIPOS DE PRODUCTO ENCONTRADOS:", ""],
    ]
    
    # Contar tipos de producto
    product_types = {}
    for idx, row in df.iterrows():
        name = str(row.get('name', '')).strip()
        part_number = str(row.get('part_number', row.get('sku', ''))).strip()
        enriched_data = enricher.enrich_with_ai(name, part_number, float(row.get('price', 0)))
        product_type = enriched_data.get('product_type', 'refaccion')
        product_types[product_type] = product_types.get(product_type, 0) + 1
    
    for product_type, count in sorted(product_types.items()):
        summary_data.append(["", f"  • {product_type}: {count} productos"])
    
    for row_idx, (label, value) in enumerate(summary_data, start=1):
        ws_summary.cell(row=row_idx, column=1, value=label)
        ws_summary.cell(row=row_idx, column=2, value=value)
        if row_idx == 1:
            ws_summary.cell(row=row_idx, column=1).font = Font(bold=True, size=14)
    
    # Guardar archivo
    wb.save(output_path)
    print(f"\n✅ Excel enriquecido guardado en: {output_path}")
    print(f"   - Total de productos: {total_products}")
    print(f"   - Tipos de producto: {len(product_types)}")


def main():
    parser = argparse.ArgumentParser(
        description='Enriquece productos con IA usando información de internet'
    )
    parser.add_argument(
        '--input',
        '-i',
        required=True,
        help='Ruta al archivo Excel de entrada (debe tener: nombre, precio, y opcionalmente número de parte, existencia)'
    )
    parser.add_argument(
        '--output',
        '-o',
        default='productos_enriquecidos.xlsx',
        help='Ruta al archivo Excel de salida (default: productos_enriquecidos.xlsx)'
    )
    parser.add_argument(
        '--openai-key',
        help='API Key de OpenAI (o usa variable de entorno OPENAI_API_KEY)'
    )
    parser.add_argument(
        '--unsplash-key',
        help='API Key de Unsplash para búsqueda de imágenes (o usa variable de entorno UNSPLASH_ACCESS_KEY)'
    )
    parser.add_argument(
        '--no-ai',
        action='store_true',
        help='No usar IA, solo enriquecimiento básico basado en palabras clave'
    )
    parser.add_argument(
        '--no-images',
        action='store_true',
        help='No buscar imágenes automáticamente'
    )
    
    args = parser.parse_args()
    
    # Validar archivo de entrada
    if not os.path.exists(args.input):
        print(f"❌ Error: El archivo {args.input} no existe")
        sys.exit(1)
    
    # Inicializar enricher
    if args.no_ai:
        print("ℹ️  Modo básico: No se usará IA")
        enricher = ProductEnricher(openai_api_key=None, unsplash_api_key=None if args.no_images else args.unsplash_key)
    else:
        enricher = ProductEnricher(openai_api_key=args.openai_key, unsplash_api_key=None if args.no_images else args.unsplash_key)
        if not enricher.openai_client:
            print("⚠️  OpenAI no está disponible. Usando modo básico.")
            print("   Para usar IA, configura OPENAI_API_KEY o usa --openai-key")
    
    if not args.no_images and not enricher.unsplash_api_key:
        print("ℹ️  Búsqueda de imágenes deshabilitada (no hay API key de Unsplash)")
        print("   Para buscar imágenes, configura UNSPLASH_ACCESS_KEY o usa --unsplash-key")
    elif not args.no_images:
        print("✅ Búsqueda de imágenes habilitada")
    
    try:
        # Leer Excel de entrada
        print(f"📖 Leyendo archivo: {args.input}")
        df = read_input_excel(args.input)
        print(f"✅ Archivo leído: {len(df)} productos encontrados")
        
        # Crear Excel enriquecido
        create_enriched_excel(df, enricher, args.output, search_images=not args.no_images)
        
        print("\n🎉 Proceso completado exitosamente!")
        print(f"\n📋 Próximos pasos:")
        print(f"   1. Revisa el archivo {args.output}")
        print(f"   2. Completa las URLs de imágenes si es necesario")
        print(f"   3. Asigna categorías usando catalogo_categorias.csv")
        print(f"   4. Importa el archivo usando el sistema de carga masiva")
        
    except Exception as e:
        print(f"❌ Error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()

